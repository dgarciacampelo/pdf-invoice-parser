import asyncio
import os
import pdfplumber
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from traceback import format_exc as traceback_format_exc


# If true, a TXT file with the extracted text will be created for each PDF file
DUMP_TEXT: bool = False
SINGLE_DUMP: bool = True

# Define the threshold for saving data to XLSX
SAVE_THRESHOLD = 25

# Define column headers based on the extracted data keys
COLUMN_HEADERS = [
    "Datos del Titular",
    "Nº Factura",
    "Fecha desde",
    "Fecha hasta",
    "CUPS",
    "Ref. Contrato Acceso",
    "P1. Energía activa (Cantidad)",
    "P1. Energía activa (Precio/u)",
    "P1. Energía activa (Total)",
]


# * Decorator to add datetime to print
def print_decorator(func):
    def wrapper(*args, **kwargs):
        try:
            return func(datetime.now().strftime("%Y-%m-%d %H:%M:%S"), *args, **kwargs)
        except KeyboardInterrupt:
            pass
        except Exception as e:
            print(f"Exception: {e}")

    return wrapper


# Custom print function with the decorator applied, dp means decorated print
@print_decorator
def dp(*args, **kwargs):
    print(*args, **kwargs)


# TODO: Add a benchmark decorator to measure execution time.


async def process_pdf(pdf_file, found_cups, pending_cups_data, threshold):
    """Given a PDF file path, extract the data and save it to an XLSX file."""

    global SINGLE_DUMP  # * Using global to avoid UnboundLocalError

    pdf_file_path = os.path.join(os.getcwd(), pdf_file)

    try:
        text = ""
        # Open the PDF file with pdfplumber
        with pdfplumber.open(pdf_file_path) as pdf:
            num_pages = len(pdf.pages)
            dp(f"Processing {pdf_file} ({num_pages} pages)...")

            for page in pdf.pages:
                text += page.extract_text()

        # Dump the extracted text to a TXT file (replacing the file extension)
        if DUMP_TEXT or SINGLE_DUMP:
            SINGLE_DUMP = False
            txt_file_name = pdf_file.replace(".pdf", ".txt")
            txt_file_path = os.path.join(os.getcwd(), txt_file_name)
            with open(txt_file_path, "w") as txt_file:
                txt_file.write(text)

        # Initialize a dictionary to store extracted data
        extracted_data = dict()

        # Extract Datos del Titular using regex
        datos_titular_match = re.search(
            r"DATOS DEL TITULAR\s*(.*?)\s*Nº Factura", text, re.DOTALL
        )
        if datos_titular_match:
            extracted_data["Datos del Titular"] = datos_titular_match.group(1).strip()

        # Extract Nº Factura using regex
        invoice_number_match = re.search(r"Nº Factura:\s*(\w+)", text)
        if invoice_number_match:
            extracted_data["Nº Factura"] = invoice_number_match.group(1)

        # Extract Período de facturación using regex
        billing_period_match = re.search(
            r"Período de facturación:\s*([\d/]+ - [\d/]+)", text
        )
        if billing_period_match:
            extracted_data["Fecha desde"] = ""
            extracted_data["Fecha hasta"] = ""

            # extracted_data["Período de facturación"] = billing_period_match.group(1)
            dates = billing_period_match.group(1).split(" - ")
            if len(dates) == 2:
                extracted_data["Fecha desde"] = dates[0]
                extracted_data["Fecha hasta"] = dates[1]

        # Extract CUPS and Ref. Contrato Acceso using regex
        cups_match = re.search(
            r"CUPS:\s*(.*?)\s*Ref. Contrato Acceso:\s*(\d+)", text, re.DOTALL
        )
        if cups_match:
            extracted_data["CUPS"] = cups_match.group(1).strip()
            extracted_data["Ref. Contrato Acceso"] = cups_match.group(2).strip()

        # Extract the line starting with "P1. Energía activa"
        p1_energia_activa_line = re.search(r"P1\. Energía activa[^\n]*", text)

        if p1_energia_activa_line:
            line_text = p1_energia_activa_line.group(0)
            # Split the line by whitespace to obtain the values
            values = re.findall(r"[\d,.]+", line_text)

            if len(values) > 3:
                # values[0] should be 1 from P1.
                extracted_data["P1. Energía activa (Cantidad)"] = values[1]
                extracted_data["P1. Energía activa (Precio/u)"] = values[2]
                extracted_data["P1. Energía activa (Total)"] = values[3]

        # Check if CUPS has been processed
        actual_cups: str = extracted_data["CUPS"]
        actual_invoice: str = extracted_data["Nº Factura"]

        if actual_cups not in found_cups:
            found_cups[actual_cups] = set()
            pending_cups_data[actual_cups] = []
            # Create a new XLSX file for this CUPS
            create_xlsx(actual_cups, found_cups)

        elif actual_invoice in found_cups[actual_cups]:
            dp(f"Skipping {pdf_file} because it has already been processed.")
            return

        # Register the invoice and link the extracted data to its CUPS code
        found_cups[actual_cups].add(actual_invoice)
        pending_cups_data[actual_cups].append(extracted_data)

        # Check if the threshold is reached
        if threshold and len(pending_cups_data[actual_cups]) >= threshold:
            # Save data to the XLSX file
            save_to_xlsx(actual_cups, pending_cups_data)

    except Exception as e:
        dp(f"Error processing {pdf_file}: {str(e)}")
        print(traceback_format_exc())


def create_xlsx(cups, found_cups):
    xlsx_file_path = os.path.join(os.getcwd(), f"{cups}.xlsx")

    # Create a new workbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(COLUMN_HEADERS)

    # Save the workbook
    workbook.save(xlsx_file_path)

    return worksheet


def save_to_xlsx(cups, pending_cups_data):
    xlsx_name = f"{cups}.xlsx"
    xlsx_file_path = os.path.join(os.getcwd(), xlsx_name)

    # Load the workbook linked to this CUPS
    workbook = load_workbook(xlsx_file_path)
    worksheet = workbook.active

    # Append rows from pending_cups_data
    for row in pending_cups_data[cups]:
        # dp(f"Row to append for {cups}: {row}")
        data_row = [row.get(key, "") for key in COLUMN_HEADERS]
        worksheet.append(data_row)

    # Save the workbook
    workbook.save(xlsx_file_path)


async def main():
    # Set variable to track processed CUPS codes values
    """
    A single CUPS code can have multiple invoices (invoice numbers). To avoid
    processing the same invoice number more than once, we need to keep track of
    the processed invoice numbers for each CUPS code. This can be done by
    refactoring the found_cups set to a dictionary, where the key is the CUPS
    code and the value is a set of processed invoice numbers.
    """
    found_cups: dict[str, set[str]] = dict()

    # Dictionary to store data pending to be saved to XLSX files, by CUPS code
    pending_cups_data: dict[str, list[dict]] = dict()

    # List all PDF files in the current directory
    pdf_files = [file for file in os.listdir(os.getcwd()) if file.endswith(".pdf")]

    for pdf_file in pdf_files:
        await process_pdf(pdf_file, found_cups, pending_cups_data, SAVE_THRESHOLD)

    # Save any remaining pending data to XLSX files
    for cups in list(found_cups):
        save_to_xlsx(cups, pending_cups_data)


if __name__ == "__main__":
    asyncio.run(main())
