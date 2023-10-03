import os
import pdfplumber
import re
from openpyxl import Workbook

# Global set variable to track processed CUPS values
processed_cups = set()

# Define the threshold for saving data to XLSX
SAVE_THRESHOLD = 25

# Define column headers based on the extracted data keys
COLUMN_HEADERS = ["Datos del Titular", "Nº Factura", "Período de facturación", "CUPS", "Ref. Contrato Acceso", "P1. Energía activa (Cantidad)", "P1. Energía activa (Precio/u)", "P1. Energía activa (Total)"]
    

async def process_pdf(pdf_file, worksheet, threshold):
    pdf_file_path = os.path.join(os.getcwd(), pdf_file)
    
    try:
        text = ''
        # Open the PDF file with pdfplumber
        with pdfplumber.open(pdf_file_path) as pdf:
            num_pages = len(pdf.pages)
            print(f"Processing {pdf_file} ({num_pages} pages)...")
            
            for page in pdf.pages:
                text += page.extract_text()
                
        # Initialize a dictionary to store extracted data
        extracted_data = dict()
        
        # Extract Datos del Titular using regex
        datos_titular_match = re.search(r'DATOS DEL TITULAR\s*(.*?)\s*Nº Factura', text, re.DOTALL)
        if datos_titular_match:
            extracted_data['Datos del Titular'] = datos_titular_match.group(1).strip()

        # Extract Nº Factura using regex
        invoice_number_match = re.search(r'Nº Factura:\s*(\w+)', text)
        if invoice_number_match:
            extracted_data['Nº Factura'] = invoice_number_match.group(1)

        # Extract Período de facturación using regex
        billing_period_match = re.search(r'Período de facturación:\s*([\d/]+ - [\d/]+)', text)
        if billing_period_match:
            extracted_data['Período de facturación'] = billing_period_match.group(1)

        # Extract CUPS and Ref. Contrato Acceso using regex
        cups_match = re.search(r'CUPS:\s*(.*?)\s*Ref. Contrato Acceso:\s*(\d+)', text, re.DOTALL)
        if cups_match:
            extracted_data['CUPS'] = cups_match.group(1).strip()
            extracted_data['Ref. Contrato Acceso'] = cups_match.group(2).strip()

        # Extract the line starting with "P1. Energía activa"
        p1_energia_activa_line = re.search(r'P1\. Energía activa[^\n]*', text)

        if p1_energia_activa_line:
            line_text = p1_energia_activa_line.group(0)
            # Split the line by whitespace to obtain the values
            values = re.findall(r'[\d,.]+', line_text)
            
            if len(values) > 3:
                # values[0] should be 1 from P1.
                extracted_data['P1. Energía activa (Cantidad)'] = values[1]
                extracted_data['P1. Energía activa (Precio/u)'] = values[2]
                extracted_data['P1. Energía activa (Total)'] = values[3]

        # Check if CUPS has been processed
        if extracted_data['CUPS'] not in processed_cups:
            # Create a new XLSX file for this CUPS
            create_xlsx(extracted_data['CUPS'], worksheet)
        
        # Append the extracted data to the worksheet
        data_row = [extracted_data.get(key, '') for key in COLUMN_HEADERS]
        worksheet.append(data_row)

        # Check if the threshold is reached
        if threshold and worksheet.max_row >= threshold:
            # Save data to the XLSX file
            save_to_xlsx(extracted_data['CUPS'], worksheet)
        
    except Exception as e:
        print(f"Error processing {pdf_file}: {str(e)}")

async def main():
    # Create a new workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    
    
    # Write column headers to the worksheet
    worksheet.append(COLUMN_HEADERS)
    
    # List all PDF files in the current directory
    pdf_files = [file for file in os.listdir(os.getcwd()) if file.endswith('.pdf')]
    
    for pdf_file in pdf_files:
        await process_pdf(pdf_file, worksheet, SAVE_THRESHOLD)
    
    # Save any remaining data to XLSX files
    for cups in processed_cups:
        save_to_xlsx(cups, worksheet)

def create_xlsx(cups, worksheet):
    xlsx_file_path = os.path.join(os.getcwd(), f'{cups}.xlsx')
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(COLUMN_HEADERS)
    workbook.save(xlsx_file_path)
    processed_cups.add(cups)

def save_to_xlsx(cups, worksheet):
    xlsx_file_path = os.path.join(os.getcwd(), f'{cups}.xlsx')
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(COLUMN_HEADERS)
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        data_row = [cell.value for cell in row]
        if data_row[3] == cups:
            worksheet.delete_rows(row[0].row)
            workbook.save(xlsx_file_path)
    processed_cups.remove(cups)

if __name__ == "__main__":
    import asyncio
    asyncio.run(main())
